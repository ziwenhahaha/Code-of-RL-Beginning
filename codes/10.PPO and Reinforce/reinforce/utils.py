import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Alignment
from PIL import Image as PILImage

def write_excel(hyperparameters ,img_path, bias, xlsx_path, 
                highlights=[], gap = 1):
    """
    :param hyperparameters: 一个列表, 表示超参数是什么
    :param img_path: 要加入进去的图片的位置
    :param bias: 插入第几个模块,以模块为单位进行偏差，也就是索引
    :param xlsx_path: 保存进入那个excel文件
    :param highlights: 哪个超参数需要高亮，是一个字符串列表
    :param gap: 每个模块相差多少行
    :return: 无
    """
    # gap是隔多少行
    hyperparameters_cnt = len(hyperparameters) # 这个加一是为了隔一格,好看一点
    # 创建一个新的工作簿

    try:
        # 尝试打开现有的Excel文件
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
    except FileNotFoundError:
        print(f"The file {xlsx_path} does not exist. Creating a new workbook.")
        wb = openpyxl.Workbook()
        ws = wb.active

    # 行偏置的和
    row_bias = bias * (hyperparameters_cnt + gap)
    #在第一列插入超参数
    row_num = row_bias+1
    for key, value in hyperparameters.items():
        cell = ws.cell(row=row_num, column=1, value=f"{key} = {value}")
        cell.alignment = Alignment(vertical='center')  # 设置纵向居中
        if key in highlights:
            cell.fill = PatternFill(start_color='F5C142', end_color='F5C142', fill_type='solid')  # 高亮黄色
        row_num += 1

    column_letter = openpyxl.utils.get_column_letter(1)
    ws.column_dimensions[column_letter].width = 30

    # 合并单元格,插入图片,设置大小,
    # 合并单元格
    column_letter = openpyxl.utils.get_column_letter(3)
    ws.merge_cells(f'{column_letter}{row_bias+1}:{column_letter}{row_bias+hyperparameters_cnt}')

    # 加载图片
    img = PILImage.open(img_path)
    img_width, img_height = img.size
    # 第3行用来
    ws.column_dimensions[column_letter].width = img_width / 1 / 8


    for row in range(row_bias+1,row_bias+len(hyperparameters)+1):
        ws.row_dimensions[row].height = img_height / hyperparameters_cnt /1.33

    img = Image(img_path)
    img.anchor = f'{column_letter}{1+row_bias}'
    ws.add_image(img)

    # 保存工作簿
    wb.save(xlsx_path)